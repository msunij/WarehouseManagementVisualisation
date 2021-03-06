# -*- coding: utf-8 -*-
"""
Created on Fri Jun 20 20:35:50 2018

@author: msunij

"""

import openpyxl
from utils import *
import time
import queue
import threading
import copy
from bresenham import bresenham
from random import choice
        
robotSpeed = 2

#initializing the warehouse data point
sideage = 3
rows = 5
cols = 5

warehouseWidth = cols+2*sideage-1 #No use so far for this variable
warehouseHeight = rows+2*sideage-1

class Item:
    def __init__(self,itemName,location,stockLevel=5,deliveryPt=0):
        self.name = itemName
        self.location = location
        self.stock = stockLevel
        self.deliveryPt = deliveryPt
        
    def addStock(self,quantity):
        self.stock += quantity
    
    def removeStock(self,quantity=1):
        self.stock -= quantity

itemQ = queue.Queue(maxsize=0)
def readExcel():
        
    D = dict()
    #fileName = input("Enter the name of the excel file: ")
    fileName = 'warehouseData.xlsx'
    workbook = openpyxl.load_workbook(fileName, data_only=True)
    sheet = workbook.active
    for i in range(2,sheet.max_row+1):
        code = sheet.cell(row=i, column=2).value
        x = sheet.cell(row=i, column=3).value
        y = sheet.cell(row=i, column=4).value
        name = sheet.cell(row=i, column=5).value
        stock = sheet.cell(row=i, column=6).value
        deliveryPt = sheet.cell(row=i, column=7).value
        
        D[code] = Item(name,scale([x,y]),stock,deliveryPt)
        itemQ.put(code)
        
    return D
    
  
#initializing the entry and exit points
class ExitPoint:
    def __init__(self,location):
        self.location = scale(location)
        
        
pointLocations = []
for i in range(sideage, sideage+cols):
    if i%2 == 1:
        pointLocations.append(ExitPoint([i, 0]))
        #print(scale([i,0]))
    else:
        pointLocations.append(ExitPoint([i, warehouseHeight]))
        #print(scale([i,warehouseHeight]))


#Creating the robot class
class Robot:
    def __init__(self, robotNumber):
        self.pos = copy.deepcopy(pointLocations[robotNumber].location) #Assigning corresponding robot to points
        #print(id(self.pos),id(pointLocations[robotNumber].location))
        self.avail = True
        self.name = 'Robot'+str(robotNumber+1)
        self.speed = 3#number of pixels per second
    
    #find the total distance robot has to travel to deliver the product
    #argument is a list of product code and delivery location
    def dist2item(self,itemCode):
        return distance(self.pos, itemDict[itemCode].location)
    
    def dist2exit(self,itemCode):
        pt = itemDict[itemCode].deliveryPt
        return distance(itemDict[itemCode].location,pointLocations[pt].location)

    def distanceCalculator(self,itemCode):
        return self.dist2item(itemCode) + self.dist2exit(itemCode)
    
    def deliver(self,itemCode):
        if itemDict[itemCode].stock < 1:
            print("Stock depleted")
        else:
            self.avail = False
            itemLoc = itemDict[itemCode].location
            exitPt = itemDict[itemCode].deliveryPt
            #print("exitPt:",exitPt)
            deliveryLoc = pointLocations[exitPt].location
            #print("deliveryLocation:",deliveryLoc)
            self.move2location(itemLoc)
            #print("item location reached")
            #print("item location:{}".format(itemLoc))
            time.sleep(1)
            itemDict[itemCode].removeStock()
            self.move2location(deliveryLoc)
            #print("deliveryPt:{}".format(exitLoc))
            time.sleep(1)
            self.avail = True
        
        
    def move2location(self,location):
        bresenhamGenerator = bresenham(self.pos[0],self.pos[1],location[0],location[1])
        x,y = next(bresenhamGenerator)
        while not self.pos == location :    
            #print(not self.pos[0]==location[0])
            #print(displacement(self.pos,location))
            x,y = next(bresenhamGenerator)
            self.pos[0] = x
            self.pos[1] = y
            time.sleep(0.01)


def toExcel(inputDict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Sl No.'
    ws['B1'] = 'ProductCode'
    ws['C1'] = 'RobotEngaged'
    ws['D1'] = 'TimeTaken'
    ws['E1'] = 'DistanceCovered'
    i = 2
    for key in inputDict.keys():  
        robotIndex, dist = closestRobotFinder(key)
        ws.cell(row=i,column=1).value = i-1
        ws.cell(row=i,column=2).value = itemDict[key].name
        ws.cell(row=i,column=3).value = robotList[robotIndex].name
        ws.cell(row=i,column=4).value = round(dist/robotSpeed,2)
        ws.cell(row=i,column=5).value = dist
        i += 1
    wb.save('output.xlsx')
    
def printAll(inputDict):
    for key in inputDict.keys():
        
        closestRobotFinderPrint(key)

def closestRobotFinder(itemCode):
    closestDist = 10000 #Maximum distance that is possible
    closestIndex = 0
    for i in range(robotCount):
        if robotList[i].avail:
            dist = robotList[i].distanceCalculator(itemCode)
            if dist < closestDist:
                closestDist = dist
                closestIndex = i
    return [closestIndex, closestDist]

def closestRobotFinderPrint(itemCode):
    with printLock:
        robotIndex, dist = closestRobotFinder(itemCode)
        #print("Product Retrieved: {}".format(itemCode))
        #print("Robot Engaged: {}".format(robotList[robotIndex].name))
        #print("from:{} item:{} final:{}".format(robotList[robotIndex].pos,itemDict[itemCode].location,pointLocations[itemDict[itemCode].deliveryPt].location))
        #robotList[robotIndex].deliver(itemCode)
    #print("Distance Covered: {}meters".format(dist))
    #print("Time Taken: {}seconds".format(round(dist/robotSpeed,2)))
    print("*********************************")

def startWork(itemCode):
    robotIndex, dist = closestRobotFinder(itemCode)
    robotList[robotIndex].deliver(itemCode)
   
    
class WorkThread(threading.Thread):
    def __init__(self,que):
        super().__init__()
        self.que = que
        
    def run(self):
        timer = [i for i in range(1,15)]
        while True:
            wait = choice(timer)
            itemCode = self.que.get()
            time.sleep(wait)
            #print(wait)
            startWork(itemCode)
            self.que.task_done()

itemDict = readExcel()

#testing git 

def main():
    for i in range(robotCount):
        thrd = WorkThread(itemQ)
        thrd.daemon = True
        thrd.start()


robotCount = 4

robotList = [ Robot(i) for i in range(robotCount)]

if __name__ == "__main__":
    main()