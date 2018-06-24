# -*- coding: utf-8 -*-
"""
Created on Fri Jun 22 20:35:50 2018

@author: msunij

"""

import string

#Utility Functions

#this function needs to be edited
def distanceMatrix():
    for letter in string.ascii_uppercase[:9]:
        print(letter,end=" :")
        for i in range(3):
            print(robotList[i].distanceDict[letter],end=",")
        print(" ")

#find the current positions of the robots 
def robotsPositions():
    for rob in robotList:
        print(rob.name, rob.pos)
        
        
robotSpeed = 1.5

#initializing the warehouse data point
warehouse = {}
sideage = 3
rows = 5
cols = 5

warehouseWidth = cols+2*sideage-1 #No use so far for this variable
warehouseHeight = rows+2*sideage-1

k = 0
for i in range(sideage, sideage+rows):
    for j in range(sideage, sideage+cols):
        warehouse[string.ascii_uppercase[k]] = [j,i]        
        k += 1
     
#initializing the entry and exit points
pointCount = 4
pointLocations = []
for i in range(sideage, sideage+cols):
    if i%2 == 1:
        pointLocations.append([i, 0])
    else:
        pointLocations.append([i, warehouseHeight])

#Creating the robot class
class robot:
    def __init__(self, robotNumber):
        self.pos = pointLocations[robotNumber] #Assigning corresponding robot to points
        self.avail = True
        self.name = 'Robot'+str(robotNumber+1)
    
    #find the total distance robot has to travel to deliver the product
    #argument is a list of product code and delivery location
    def distanceCalculator(self,item):
        dist2item = (abs(self.pos[0]-warehouse[item[0]][0])
                    +abs(self.pos[1]-warehouse[item[0]][1]))
        dist2exit = (abs(pointLocations[item[1]][0]-warehouse[item[0]][0])
                    +abs(pointLocations[item[1]][1]-warehouse[item[0]][1]))
        distance = dist2item+dist2exit
        return distance

robotCount = 3

robotList = [ robot(i) for i in range(robotCount)]

def allWarehouse(inputDict):
    for key, value in inputDict.items():
        
        closestRobotFinderPrint([key, value[2]])

def closestRobotFinder(item):
    closestDist = 10000 #Maximum distance that is possible
    closestIndex = 0
    for i in range(robotCount):
        if robotList[i].avail:
            dist = robotList[i].distanceCalculator(item)
            if dist < closestDist:
                closestDist = dist
                closestIndex = i
        
    robot2use = robotList[closestIndex]
    robot2use.pos = pointLocations[item[1]]
    return [robot2use.name, closestDist]

def closestRobotFinderPrint(item):
	robotNumber, timeTaken = closestRobotFinder(item)
	print("Robot Engaged: {}".format(robotNumber))
	print("Distance Covered: {}meters".format(timeTaken))
	print("Time Taken: {}seconds".format(timeTaken*robotSpeed))
	  
def mainHardCoded():
	item2retrive = ['E', 0]
	closestRobotFinderPrint(item2retrive)

def inputAsker():
    name = input('Item Code :')
    loc = int(input('Exit point :'))
    return [name, loc]

def mainLooped():
    while True:
        name, loc = inputAsker()
        closestRobotFinderPrint([name,loc])
        check = input("Check another item?(y/n): ")
        if check == 'n':
            break
        
def readExcel():
    
    import openpyxl
    from collections import OrderedDict
    
    warehouse = OrderedDict()
    fileName = input("Enter the name of the excel file: ")
    workbook = openpyxl.load_workbook(fileName, data_only=True)
    sheet = workbook.active
    for i in range(2,sheet.max_row+1):
        code = sheet.cell(row=i, column=1).value
        x = sheet.cell(row=i, column=2).value
        y = sheet.cell(row=i, column=3).value
        pt = sheet.cell(row=i, column=4).value
        warehouse.update({code:[x,y,pt]})
    return warehouse
    
        
def mainFromExcel():
    warehouse = readExcel()
    allWarehouse(warehouse)

def main():
	res = input("How do you want to proceed?\n
			1. From and excel file\n
		  	2. Manual input of Product Code[A - Y] and\n
				delivery location[0-3]")
	if res == '1':
		mainLooped()
	else:
		mainFromExcel()

main()
